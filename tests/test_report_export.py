# Copyright (c) Microsoft Corporation.
# Licensed under the MIT license.

"""Tests for the PDF and Excel report exporter module."""

from __future__ import annotations

import io

from azure_sub_migrator.report_export import generate_excel, generate_pdf

# ── Shared fixtures ───────────────────────────────────────────────────

_SAMPLE_SCAN = {
    "transfer_safe": [
        {
            "name": "vm-web-01",
            "type": "Microsoft.Compute/virtualMachines",
            "resource_group": "rg-prod",
            "location": "eastus",
        },
        {
            "name": "nsg-web",
            "type": "Microsoft.Network/networkSecurityGroups",
            "resource_group": "rg-prod",
            "location": "eastus",
        },
    ],
    "requires_action": [
        {
            "name": "kv-prod",
            "type": "Microsoft.KeyVault/vaults",
            "resource_group": "rg-prod",
            "location": "eastus",
            "timing": "both",
            "pre_action": "Remove all access policies before transfer.",
            "post_action": "Re-add access policies after transfer.",
            "doc_url": "https://learn.microsoft.com/en-us/azure/key-vault/general/move-subscription",
        },
        {
            "name": "sql-prod",
            "type": "Microsoft.Sql/servers",
            "resource_group": "rg-data",
            "location": "westus",
            "timing": "pre",
            "pre_action": "Disable Entra authentication.",
            "post_action": "",
            "doc_url": "https://learn.microsoft.com/en-us/azure/azure-sql/database/move-resources-across-regions",
        },
    ],
}

_EMPTY_SCAN = {
    "transfer_safe": [],
    "requires_action": [],
}


# ── PDF tests ─────────────────────────────────────────────────────────

class TestGeneratePDF:
    def test_returns_bytes(self):
        result = generate_pdf(_SAMPLE_SCAN, subscription_id="sub-123")
        assert isinstance(result, (bytes, bytearray))
        assert len(result) > 0

    def test_starts_with_pdf_header(self):
        result = generate_pdf(_SAMPLE_SCAN, subscription_id="sub-123")
        assert result[:5] == b"%PDF-"

    def test_empty_scan_produces_valid_pdf(self):
        result = generate_pdf(_EMPTY_SCAN, subscription_id="sub-empty")
        assert result[:5] == b"%PDF-"
        assert len(result) > 100

    def test_no_subscription_id(self):
        """Works without a subscription ID."""
        result = generate_pdf(_SAMPLE_SCAN)
        assert result[:5] == b"%PDF-"

    def test_pdf_has_multiple_pages(self):
        """Report with data should produce a multi-page PDF."""
        result = generate_pdf(_SAMPLE_SCAN, subscription_id="sub-123")
        # fpdf2 writes /Count N for the number of pages
        pdf_text = bytes(result).decode("latin-1", errors="ignore")
        assert "/Count 5" in pdf_text  # cover + safe + action + steps + footer renders


# ── Excel tests ───────────────────────────────────────────────────────

class TestGenerateExcel:
    def test_returns_bytes(self):
        result = generate_excel(_SAMPLE_SCAN, subscription_id="sub-123")
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_xlsx_magic_bytes(self):
        """XLSX files are ZIP archives — start with PK signature."""
        result = generate_excel(_SAMPLE_SCAN, subscription_id="sub-123")
        assert result[:2] == b"PK"

    def test_has_three_sheets(self):
        from openpyxl import load_workbook

        result = generate_excel(_SAMPLE_SCAN, subscription_id="sub-123")
        wb = load_workbook(io.BytesIO(result))
        assert len(wb.sheetnames) == 3
        assert "Summary" in wb.sheetnames
        assert "Transfer-Safe" in wb.sheetnames
        assert "Requires Action" in wb.sheetnames

    def test_transfer_safe_data_rows(self):
        from openpyxl import load_workbook

        result = generate_excel(_SAMPLE_SCAN, subscription_id="sub-123")
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Transfer-Safe"]
        # Header + 2 data rows
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 2
        names = [row[0] for row in data_rows]
        assert "vm-web-01" in names
        assert "nsg-web" in names

    def test_requires_action_data_rows(self):
        from openpyxl import load_workbook

        result = generate_excel(_SAMPLE_SCAN, subscription_id="sub-123")
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Requires Action"]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 2
        names = [row[0] for row in data_rows]
        assert "kv-prod" in names
        assert "sql-prod" in names
        # Check timing column
        timings = [row[3] for row in data_rows]
        assert "BOTH" in timings
        assert "PRE" in timings

    def test_empty_scan_produces_valid_xlsx(self):
        result = generate_excel(_EMPTY_SCAN, subscription_id="sub-empty")
        assert result[:2] == b"PK"

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(result))
        ws = wb["Transfer-Safe"]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 0

    def test_summary_sheet_counts(self):
        from openpyxl import load_workbook

        result = generate_excel(_SAMPLE_SCAN, subscription_id="sub-123")
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Summary"]
        # Row 7 = Total, Row 8 = Transfer-Safe, Row 9 = Requires Action
        assert ws["B7"].value == 4   # total
        assert ws["B8"].value == 2   # transfer-safe
        assert ws["B9"].value == 2   # requires action
