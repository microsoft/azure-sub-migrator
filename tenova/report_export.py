"""PDF and Excel report exporter for migration scan results.

Generates professional-looking reports that can be shared with
stakeholders, change-advisory boards, or attached to migration tickets.

Supported formats:
- **PDF** — branded, multi-page report using ``fpdf2``
- **Excel** — multi-sheet workbook using ``openpyxl``
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

from tenova.logger import get_logger

logger = get_logger("report_export")


# ──────────────────────────────────────────────────────────────────────
# PDF export
# ──────────────────────────────────────────────────────────────────────

def generate_pdf(
    scan_result: dict[str, Any],
    subscription_id: str = "",
) -> bytes:
    """Generate a PDF migration report and return it as bytes.

    Parameters
    ----------
    scan_result:
        The scan result dict with ``transfer_safe`` and ``requires_action`` lists.
    subscription_id:
        Azure subscription ID (shown in the header).

    Returns
    -------
    PDF file contents as bytes.
    """
    from fpdf import FPDF

    transfer_safe = scan_result.get("transfer_safe", [])
    requires_action = scan_result.get("requires_action", [])
    # Count children nested inside parent entries
    child_count = sum(len(r.get("children", [])) for r in requires_action)
    total = len(transfer_safe) + len(requires_action) + child_count
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)

    # ── Cover / Title ──────────────────────────────────────────────
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 24)
    pdf.cell(0, 20, "Tenova", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.cell(0, 12, "Migration Report", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(10)

    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 8, f"Generated: {generated}", new_x="LMARGIN", new_y="NEXT", align="C")
    if subscription_id:
        pdf.cell(0, 8, f"Subscription: {subscription_id}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(10)

    # ── Summary box ────────────────────────────────────────────────
    pdf.set_fill_color(240, 240, 245)
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 12, "  Summary", new_x="LMARGIN", new_y="NEXT", fill=True)
    pdf.ln(3)

    _summary_row(pdf, "Total Resources", str(total))
    _summary_row(pdf, "Transfer-Safe", str(len(transfer_safe)), (34, 139, 34))
    _summary_row(pdf, "Requires Action", str(len(requires_action) + child_count), (200, 50, 50))
    pdf.ln(5)

    pdf.set_font("Helvetica", "I", 9)
    pdf.set_text_color(100, 100, 100)
    pdf.multi_cell(
        0, 5,
        _safe(
            "All resources physically move with the subscription. 'Requires Action' resources "
            "have tenant-bound dependencies that need preparation before the transfer and/or "
            "reconfiguration after."
        ),
    )
    pdf.set_text_color(0, 0, 0)

    # ── Transfer-Safe table ────────────────────────────────────────
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_fill_color(34, 139, 34)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 10, f"  Transfer-Safe Resources ({len(transfer_safe)})", new_x="LMARGIN", new_y="NEXT", fill=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(3)

    if transfer_safe:
        col_widths = [60, 50, 40, 40]
        headers = ["Name", "Type", "Resource Group", "Location"]
        _table_header(pdf, headers, col_widths, (34, 139, 34))

        pdf.set_font("Helvetica", "", 8)
        for idx, r in enumerate(transfer_safe):
            fill = idx % 2 == 0
            if fill:
                pdf.set_fill_color(245, 250, 245)
            _table_row(pdf, [
                r.get("name", ""),
                _short_type(r.get("type", "")),
                r.get("resource_group", ""),
                r.get("location", ""),
            ], col_widths, fill)
    else:
        pdf.set_font("Helvetica", "I", 10)
        pdf.cell(0, 10, "No transfer-safe resources found.", new_x="LMARGIN", new_y="NEXT")

    # ── Requires-Action table ──────────────────────────────────────
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_fill_color(200, 50, 50)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 10, f"  Requires-Action Resources ({len(requires_action)})", new_x="LMARGIN", new_y="NEXT", fill=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(3)

    if requires_action:
        col_widths = [35, 40, 15, 45, 45]
        headers = ["Name", "Type", "Timing", "Pre-Transfer Action", "Post-Transfer Action"]
        _table_header(pdf, headers, col_widths, (200, 50, 50))

        pdf.set_font("Helvetica", "", 7)
        row_num = 0
        for r in requires_action:
            fill = row_num % 2 == 0
            if fill:
                pdf.set_fill_color(255, 245, 245)
            timing = (r.get("timing", "post") or "post").upper()
            name = r.get("name", "")
            children = r.get("children", [])
            if children:
                name = f"{name} (+{len(children)} sub-resource(s))"
            _table_row(pdf, [
                name,
                _short_type(r.get("type", "")),
                timing,
                r.get("pre_action", "") or "-",
                r.get("post_action", "") or "-",
            ], col_widths, fill)
            row_num += 1
            # Render child rows indented
            for child in children:
                fill = row_num % 2 == 0
                if fill:
                    pdf.set_fill_color(255, 250, 250)
                c_timing = (child.get("timing", "post") or "post").upper()
                _table_row(pdf, [
                    f"  >> {child.get('name', '')}",
                    _short_type(child.get("type", "")),
                    c_timing,
                    child.get("pre_action", "") or "-",
                    child.get("post_action", "") or "-",
                ], col_widths, fill)
                row_num += 1
    else:
        pdf.set_font("Helvetica", "I", 10)
        pdf.cell(0, 10, "No resources requiring action found.", new_x="LMARGIN", new_y="NEXT")

    # ── Next Steps ─────────────────────────────────────────────────
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_fill_color(52, 73, 94)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 10, "  Next Steps", new_x="LMARGIN", new_y="NEXT", fill=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(5)

    steps = [
        "Review the transfer-safe vs. requires-action classification above.",
        "Complete ALL Pre-Transfer Actions listed above (critical items first).",
        "Export RBAC role assignments and custom roles (permanently deleted during transfer).",
        "Back up Key Vault secrets, certificates, and keys.",
        "Disable Entra auth on SQL/MySQL/PostgreSQL if applicable.",
        "Export Azure Policy definitions and assignments.",
        "Initiate the subscription transfer via Azure Portal or CLI.",
        "Complete ALL Post-Transfer Actions listed above.",
        "Recreate role assignments and managed identities in the target tenant.",
        "Validate all services in the target tenant.",
    ]
    pdf.set_font("Helvetica", "", 10)
    for i, step in enumerate(steps, 1):
        pdf.multi_cell(0, 7, _safe(f"  {i}. {step}"))
        pdf.ln(1)

    pdf.ln(5)
    pdf.set_font("Helvetica", "I", 9)
    pdf.set_text_color(0, 102, 204)
    pdf.cell(
        0, 6,
        "Reference: https://learn.microsoft.com/en-us/azure/role-based-access-control/transfer-subscription",
        new_x="LMARGIN", new_y="NEXT",
    )
    pdf.set_text_color(0, 0, 0)

    # ── Footer on every page ───────────────────────────────────────
    for page_num in range(1, pdf.pages_count + 1):
        pdf.page = page_num
        pdf.set_y(-15)
        pdf.set_font("Helvetica", "I", 8)
        pdf.set_text_color(150, 150, 150)
        pdf.cell(0, 10, _safe(f"tenova  |  Page {page_num} of {pdf.pages_count}"), align="C")
        pdf.set_text_color(0, 0, 0)

    return pdf.output()


# ──────────────────────────────────────────────────────────────────────
# Excel export
# ──────────────────────────────────────────────────────────────────────

def generate_excel(
    scan_result: dict[str, Any],
    subscription_id: str = "",
) -> bytes:
    """Generate an Excel migration report and return it as bytes.

    Creates a workbook with three sheets:
    - **Summary** — overview with counts
    - **Transfer-Safe** — all resources that transfer cleanly
    - **Requires Action** — resources needing pre/post-transfer work

    Parameters
    ----------
    scan_result:
        The scan result dict with ``transfer_safe`` and ``requires_action`` lists.
    subscription_id:
        Azure subscription ID (shown in the Summary sheet).

    Returns
    -------
    Excel file contents as bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    transfer_safe = scan_result.get("transfer_safe", [])
    requires_action = scan_result.get("requires_action", [])
    # Count children nested inside parent entries
    child_count = sum(len(r.get("children", [])) for r in requires_action)
    total = len(transfer_safe) + len(requires_action) + child_count
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    wb = Workbook()
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    wrap_align = Alignment(wrap_text=True, vertical="top")

    # ── Summary sheet ──────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.sheet_properties.tabColor = "4472C4"

    ws_summary["A1"] = "Tenova Migration Report"
    ws_summary["A1"].font = Font(bold=True, size=16, color="2E4057")
    ws_summary.merge_cells("A1:D1")

    ws_summary["A3"] = "Generated"
    ws_summary["B3"] = generated
    ws_summary["A4"] = "Subscription ID"
    ws_summary["B4"] = subscription_id or "N/A"
    for r in range(3, 5):
        ws_summary[f"A{r}"].font = Font(bold=True)

    summary_header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    ws_summary["A6"] = "Category"
    ws_summary["B6"] = "Count"
    for cell in [ws_summary["A6"], ws_summary["B6"]]:
        cell.font = header_font
        cell.fill = summary_header_fill
        cell.border = thin_border

    rows = [
        ("Total Resources", total),
        ("Transfer-Safe", len(transfer_safe)),
        ("Requires Action", len(requires_action) + child_count),
    ]
    for i, (label, count) in enumerate(rows, 7):
        ws_summary[f"A{i}"] = label
        ws_summary[f"B{i}"] = count
        ws_summary[f"A{i}"].border = thin_border
        ws_summary[f"B{i}"].border = thin_border
        if label == "Requires Action" and count > 0:
            ws_summary[f"B{i}"].font = Font(bold=True, color="CC3333")

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 40

    # ── Transfer-Safe sheet ────────────────────────────────────────
    ws_safe = wb.create_sheet("Transfer-Safe")
    ws_safe.sheet_properties.tabColor = "228B22"

    safe_headers = ["Name", "Type", "Resource Group", "Location"]
    safe_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
    for col_idx, header in enumerate(safe_headers, 1):
        cell = ws_safe.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = safe_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(transfer_safe, 2):
        values = [r.get("name", ""), r.get("type", ""), r.get("resource_group", ""), r.get("location", "")]
        stripe = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid") if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(values, 1):
            cell = ws_safe.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = wrap_align
            if stripe:
                cell.fill = stripe

    for col_idx in range(1, len(safe_headers) + 1):
        ws_safe.column_dimensions[get_column_letter(col_idx)].width = 35

    # ── Requires-Action sheet ──────────────────────────────────────
    ws_action = wb.create_sheet("Requires Action")
    ws_action.sheet_properties.tabColor = "CC3333"

    action_headers = [
        "Name", "Type", "Resource Group", "Timing",
        "Pre-Transfer Action", "Post-Transfer Action", "MS Learn Docs",
    ]
    action_fill = PatternFill(start_color="CC3333", end_color="CC3333", fill_type="solid")
    for col_idx, header in enumerate(action_headers, 1):
        cell = ws_action.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = action_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    excel_row = 2
    child_indent_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")
    for r in requires_action:
        timing = (r.get("timing", "post") or "post").upper()
        children = r.get("children", [])
        name = r.get("name", "")
        if children:
            name = f"{name}  (+{len(children)} sub-resource(s))"
        values = [
            name,
            r.get("type", ""),
            r.get("resource_group", ""),
            timing,
            r.get("pre_action", "") or "-",
            r.get("post_action", "") or "-",
            r.get("doc_url", "") or "-",
        ]
        stripe = (
            PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid")
            if excel_row % 2 == 0
            else None
        )
        for col_idx, val in enumerate(values, 1):
            cell = ws_action.cell(row=excel_row, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = wrap_align
            if stripe:
                cell.fill = stripe
            # Timing badge colors
            if col_idx == 4:
                if timing == "PRE":
                    cell.font = Font(bold=True, color="CC3333")
                elif timing == "BOTH":
                    cell.font = Font(bold=True, color="CC6600")
                else:
                    cell.font = Font(bold=True, color="CC9900")
            # Bold parent name if it has children
            if col_idx == 1 and children:
                cell.font = Font(bold=True)
        excel_row += 1
        # Render child rows indented
        for child in children:
            c_timing = (child.get("timing", "post") or "post").upper()
            c_values = [
                f"    ↳ {child.get('name', '')}",
                child.get("type", ""),
                child.get("resource_group", ""),
                c_timing,
                child.get("pre_action", "") or "-",
                child.get("post_action", "") or "-",
                child.get("doc_url", "") or "-",
            ]
            for col_idx, val in enumerate(c_values, 1):
                cell = ws_action.cell(row=excel_row, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = wrap_align
                cell.fill = child_indent_fill
                if col_idx == 4:
                    if c_timing == "PRE":
                        cell.font = Font(bold=True, color="CC3333")
                    elif c_timing == "BOTH":
                        cell.font = Font(bold=True, color="CC6600")
                    else:
                        cell.font = Font(bold=True, color="CC9900")
            excel_row += 1

    col_widths = [30, 35, 25, 10, 40, 40, 50]
    for col_idx, width in enumerate(col_widths, 1):
        ws_action.column_dimensions[get_column_letter(col_idx)].width = width

    # Add autofilters to data sheets
    if transfer_safe:
        ws_safe.auto_filter.ref = f"A1:{get_column_letter(len(safe_headers))}{len(transfer_safe) + 1}"
    if requires_action:
        ws_action.auto_filter.ref = f"A1:{get_column_letter(len(action_headers))}{excel_row - 1}"

    # Freeze header row on data sheets
    ws_safe.freeze_panes = "A2"
    ws_action.freeze_panes = "A2"

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────
# PDF helper functions
# ──────────────────────────────────────────────────────────────────────

# Map of common Unicode chars to ASCII-safe equivalents for Helvetica
_UNICODE_MAP: dict[str, str] = {
    "\u2192": "->",   # →
    "\u2190": "<-",   # ←
    "\u2194": "<->",  # ↔
    "\u2197": "->",   # ↗
    "\u2713": "[x]",  # ✓
    "\u2714": "[x]",  # ✔
    "\u2717": "[!]",  # ✗
    "\u2718": "[!]",  # ✘
    "\u26d4": "[!]",  # ⛔
    "\u26a0": "[!]",  # ⚠
    "\ufe0f": "",     # variation selector (invisible)
    "\u2022": "*",    # •
    "\u2013": "-",    # –
    "\u2014": "--",   # —
    "\u2018": "'",    # '
    "\u2019": "'",    # '
    "\u201c": '"',    # "
    "\u201d": '"',    # "
    "\u2026": "...",  # …
    "\u00bb": ">>",   # »
    "\u00ab": "<<",   # «
    "\u2009": " ",    # thin space
    "\u00a0": " ",    # non-breaking space
}


def _safe(text: str) -> str:
    """Replace Unicode characters unsupported by Helvetica with ASCII equivalents."""
    for char, replacement in _UNICODE_MAP.items():
        text = text.replace(char, replacement)
    # Strip any remaining non-latin-1 characters
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _summary_row(pdf: Any, label: str, value: str, color: tuple[int, int, int] | None = None) -> None:
    """Render a single summary key-value row."""
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(90, 8, f"    {label}")
    if color:
        pdf.set_text_color(*color)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, value, new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)


def _table_header(pdf: Any, headers: list[str], widths: list[int], color: tuple[int, int, int]) -> None:
    """Render a coloured table header row."""
    pdf.set_fill_color(*color)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    for header, width in zip(headers, widths):
        pdf.cell(width, 8, f" {header}", border=1, fill=True)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)


def _table_row(pdf: Any, values: list[str], widths: list[int], fill: bool = False) -> None:
    """Render a data row (supports automatic page break)."""
    # Check if we need a page break (leave room for the row + margin)
    if pdf.get_y() > pdf.h - 30:
        pdf.add_page()

    for val, width in zip(values, widths):
        text = _safe(str(val))[:60]  # Sanitize + truncate
        pdf.cell(width, 7, f" {text}", border=1, fill=fill)
    pdf.ln()


def _short_type(resource_type: str) -> str:
    """Shorten 'Microsoft.KeyVault/vaults' to 'KeyVault/vaults'."""
    if resource_type.startswith("Microsoft."):
        return resource_type[len("Microsoft."):]
    return resource_type
